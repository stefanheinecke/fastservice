import io
import os
import json
import datetime

import numpy as np
import pandas as pd
from flask import Flask, render_template, send_file, Response, request
from data import Predictor, get_symbol_name, robo_index_backtest
from sqlalchemy import create_engine, text

app = Flask(__name__)

DATABASE_URL = os.environ.get("DATABASE_URL")
SYMBOL = "GC=F"


def _get_predictor(symbol=None):
    return Predictor(DATABASE_URL, symbol or SYMBOL)


def _json_response(data):
    """Serialize to JSON, converting NaN/Infinity to null."""
    body = json.dumps(data, default=str, allow_nan=False)
    return Response(body, mimetype="application/json")

@app.route("/")
def index():
    return render_template("index.html")


# -- API --

# Restore /api/summary-predictions endpoint
@app.route("/api/summary-predictions")
def api_summary_predictions():
    try:
        engine = create_engine(DATABASE_URL)
        with engine.connect() as conn:
            symbols = pd.read_sql("SELECT DISTINCT symbol FROM predictions ORDER BY symbol", conn)["symbol"].tolist()
        results = []
        for symbol in symbols:
            try:
                predictor = _get_predictor(symbol)
                df, correct_direction_perc, mae, rmse, mape, close_correct, mae_pct, rmse_pct = predictor.fetch_prediction_history()
                forecast = predictor.fetch_next_day_forecast()
                next_pred_date = None
                next_pred_value = None
                if forecast:
                    next_pred_date = forecast.get("date")
                    next_pred_value = forecast.get("predicted_close")
                last_real_close = round(float(df.iloc[0]["Real_Close"]), 2) if not df.empty and pd.notna(df.iloc[0]["Real_Close"]) else None
                last_real_close_date = str(df.iloc[0]["Date"]) if not df.empty else None
                last_pred_close = round(float(df.iloc[0]["Predicted_Close"]), 2) if not df.empty and pd.notna(df.iloc[0]["Predicted_Close"]) else None
                results.append({
                    "symbol": symbol,
                    "name": get_symbol_name(symbol),
                    "next_pred_date": next_pred_date,
                    "next_pred_value": next_pred_value,
                    "last_real_close": last_real_close,
                    "last_real_close_date": last_real_close_date,
                    "last_pred_close": last_pred_close,
                    "mae": round(mae, 2) if mae is not None else None,
                    "mae_pct": round(mae_pct, 2) if mae_pct is not None else None,
                    "rmse": round(rmse, 2) if rmse is not None else None,
                    "rmse_pct": round(rmse_pct, 2) if rmse_pct is not None else None,
                    "mape": round(mape, 2) if mape is not None else None,
                    "correct_direction": round(correct_direction_perc, 2) if correct_direction_perc is not None else None,
                    "close_correct": round(close_correct, 2) if close_correct is not None else None,
                })
            except Exception as e:
                import traceback
                print(f"Error processing symbol {symbol}: {e}\n{traceback.format_exc()}")
        return _json_response(results)
    except Exception as e:
        import traceback
        print(f"Error in /api/summary-predictions: {e}\n{traceback.format_exc()}")
        return _json_response({"error": str(e), "trace": traceback.format_exc()}), 500

@app.route("/api/data-summary")
def api_data_summary():
    from sqlalchemy import create_engine, text
    engine = create_engine(DATABASE_URL)
    query = text("""
        SELECT symbol,
               MIN(date) AS min_date,
               MAX(date) AS max_date,
               COUNT(*) AS total_rows,
               COUNT(real_close) AS rows_with_real
        FROM predictions
        GROUP BY symbol
        ORDER BY symbol
    """)
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    rows = json.loads(df.to_json(orient="records", date_format="iso"))
    for row in rows:
        row["name"] = get_symbol_name(row["symbol"]) if row.get("symbol") else None
    return _json_response(rows)


@app.route("/api/index-components/<index_name>")
def api_index_components(index_name):
    """Return ticker symbols for a known stock index."""
    INDEX_COMPONENTS = {
        "smi": [
            "ABBN.SW", "ALC.SW", "CFR.SW", "GEBN.SW", "GIVN.SW",
            "HOLN.SW", "KNIN.SW", "LONN.SW", "NESN.SW", "NOVN.SW",
            "PGHN.SW", "ROG.SW", "SCMN.SW", "SDZ.SW", "SIKA.SW",
            "SLHN.SW", "SOON.SW", "SREN.SW", "UBSG.SW", "ZURN.SW",
        ],
    }
    name = index_name.lower()
    if name not in INDEX_COMPONENTS:
        return _json_response({"error": f"Unknown index: {index_name}"}), 404
    return _json_response({"index": index_name, "tickers": INDEX_COMPONENTS[name]})


SMI_TICKERS = [
    "ABBN.SW", "ALC.SW", "CFR.SW", "GEBN.SW", "GIVN.SW",
    "HOLN.SW", "KNIN.SW", "LONN.SW", "NESN.SW", "NOVN.SW",
    "PGHN.SW", "ROG.SW", "SCMN.SW", "SDZ.SW", "SIKA.SW",
    "SLHN.SW", "SOON.SW", "SREN.SW", "UBSG.SW", "ZURN.SW",
]


@app.route("/api/robo-index")
def api_robo_index():
    """Run the Robo-Index backtest and return results."""
    weeks = request.args.get("weeks", default=52, type=int)
    weeks = min(max(weeks, 4), 156)  # clamp 4–156 weeks
    rebal = request.args.get("rebal", default="3M", type=str)
    if rebal not in ("D", "W", "M", "3M"):
        rebal = "3M"
    try:
        result = robo_index_backtest(DATABASE_URL, SMI_TICKERS, lookback_weeks=weeks, rebal_freq=rebal)
        # Strip daily_detail from JSON response (only used for CSV download)
        json_result = {k: v for k, v in result.items() if k != "daily_detail"}
        return _json_response(json_result)
    except Exception as e:
        import traceback
        print(f"Error in /api/robo-index: {e}\n{traceback.format_exc()}")
        return _json_response({"error": str(e)}), 500


@app.route("/api/robo-index/report")
def api_robo_index_report():
    """Run the Robo-Index backtest and return a detailed Excel report with formulas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    weeks = request.args.get("weeks", default=52, type=int)
    weeks = min(max(weeks, 4), 156)
    rebal = request.args.get("rebal", default="3M", type=str)
    if rebal not in ("D", "W", "M", "3M"):
        rebal = "3M"
    try:
        result = robo_index_backtest(DATABASE_URL, SMI_TICKERS, lookback_weeks=weeks, rebal_freq=rebal)
        if "error" in result:
            return _json_response({"error": result["error"]}), 400

        detail = result.get("daily_detail", [])
        compositions = result.get("compositions", [])
        p_stats = result.get("portfolio_stats", {})
        s_stats = result.get("smi_stats", {})

        if not detail:
            return _json_response({"error": "No data for report"}), 400

        wb = Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        rebal_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        pct_fmt = '0.00%'
        num_fmt = '#,##0.00'

        def _style_header(ws, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")

        # ================================================================
        # Sheet 1 — Portfolio (one row per day)
        # ================================================================
        ws1 = wb.active
        ws1.title = "Portfolio"
        headers1 = [
            "Date", "Rebalance?",
            "Portfolio Value", "Portfolio Daily Return",
            "SMI Value", "SMI Daily Return",
            "Excess Return",
        ]
        ws1.append(headers1)
        _style_header(ws1, len(headers1))

        # Build unique dates in order
        seen_dates = []
        seen_set = set()
        date_rebal = {}
        date_port = {}
        date_smi = {}
        for d in detail:
            dt = d["date"]
            if dt not in seen_set:
                seen_dates.append(dt)
                seen_set.add(dt)
                date_rebal[dt] = d["is_rebalance_day"]
                date_port[dt] = d["portfolio_value"]
                date_smi[dt] = d["smi_value"]

        for i, dt in enumerate(seen_dates):
            row = i + 2  # 1-indexed, header is row 1
            ws1.cell(row=row, column=1, value=dt)
            ws1.cell(row=row, column=2, value="Yes" if date_rebal[dt] else "")
            ws1.cell(row=row, column=3, value=date_port[dt])
            ws1.cell(row=row, column=3).number_format = num_fmt
            if i == 0:
                ws1.cell(row=row, column=4, value=0)
            else:
                # Formula: (C_row / C_prev) - 1
                ws1.cell(row=row, column=4).value = f"=C{row}/C{row-1}-1"
            ws1.cell(row=row, column=4).number_format = pct_fmt
            ws1.cell(row=row, column=5, value=date_smi[dt])
            ws1.cell(row=row, column=5).number_format = num_fmt
            if i == 0:
                ws1.cell(row=row, column=6, value=0)
            else:
                ws1.cell(row=row, column=6).value = f"=E{row}/E{row-1}-1"
            ws1.cell(row=row, column=6).number_format = pct_fmt
            # Excess return = Portfolio return - SMI return
            ws1.cell(row=row, column=7).value = f"=D{row}-F{row}"
            ws1.cell(row=row, column=7).number_format = pct_fmt

            if date_rebal[dt]:
                for c in range(1, len(headers1) + 1):
                    ws1.cell(row=row, column=c).fill = rebal_fill

        # Summary row
        last_data = len(seen_dates) + 1
        summ_row = last_data + 2
        ws1.cell(row=summ_row, column=1, value="Total Return").font = Font(bold=True)
        ws1.cell(row=summ_row, column=3).value = f"=C{last_data}/C2-1"
        ws1.cell(row=summ_row, column=3).number_format = pct_fmt
        ws1.cell(row=summ_row, column=3).font = Font(bold=True)
        ws1.cell(row=summ_row, column=5).value = f"=E{last_data}/E2-1"
        ws1.cell(row=summ_row, column=5).number_format = pct_fmt
        ws1.cell(row=summ_row, column=5).font = Font(bold=True)

        for c in range(1, len(headers1) + 1):
            ws1.column_dimensions[get_column_letter(c)].width = 18

        # ================================================================
        # Sheet 2 — Holdings Detail (one row per stock per day)
        # ================================================================
        ws2 = wb.create_sheet("Holdings Detail")
        headers2 = [
            "Date", "Symbol", "Name", "Sector",
            "Weight %", "Dir. Accuracy %",
            "Predicted Close", "Real Close", "Market Close",
            "Pred vs Real Diff", "Pred Direction",
            "Daily Stock Return", "Weighted Return",
            "Rebalance?",
        ]
        ws2.append(headers2)
        _style_header(ws2, len(headers2))

        # Build a lookup: (date, symbol) -> previous day's market_close row number
        # so we can write return formulas referencing the previous day
        prev_row_map = {}  # (symbol) -> last row number written
        for idx, d in enumerate(detail):
            row = idx + 2
            sym = d["symbol"]
            ws2.cell(row=row, column=1, value=d["date"])
            ws2.cell(row=row, column=2, value=sym)
            ws2.cell(row=row, column=3, value=d["name"])
            ws2.cell(row=row, column=4, value=d["sector"])
            ws2.cell(row=row, column=5, value=d["weight_pct"])
            ws2.cell(row=row, column=5).number_format = '0.00'
            ws2.cell(row=row, column=6, value=d["direction_accuracy"])
            ws2.cell(row=row, column=6).number_format = '0.00'
            ws2.cell(row=row, column=7, value=d["predicted_close"])
            ws2.cell(row=row, column=7).number_format = num_fmt
            ws2.cell(row=row, column=8, value=d["real_close"])
            ws2.cell(row=row, column=8).number_format = num_fmt
            ws2.cell(row=row, column=9, value=d["market_close"])
            ws2.cell(row=row, column=9).number_format = num_fmt
            # Pred vs Real Diff = Predicted - Real (col G - col H)
            ws2.cell(row=row, column=10).value = f'=IF(AND(G{row}<>"",H{row}<>""),G{row}-H{row},"")'
            ws2.cell(row=row, column=10).number_format = num_fmt
            # Pred Direction: UP if predicted > real of prev day
            ws2.cell(row=row, column=11).value = f'=IF(G{row}<>"",IF(G{row}>H{row},"UP","DOWN"),"")'
            # Daily Stock Return = market_close_today / market_close_prev_day - 1
            if sym in prev_row_map:
                pr = prev_row_map[sym]
                ws2.cell(row=row, column=12).value = f"=IF(AND(I{row}<>\"\",I{pr}<>\"\"),I{row}/I{pr}-1,\"\")"
                ws2.cell(row=row, column=12).number_format = pct_fmt
                # Weighted Return = weight% / 100 * daily return
                ws2.cell(row=row, column=13).value = f'=IF(L{row}<>"",E{row}/100*L{row},"")'
                ws2.cell(row=row, column=13).number_format = pct_fmt
            else:
                ws2.cell(row=row, column=12, value="")
                ws2.cell(row=row, column=13, value="")
            ws2.cell(row=row, column=14, value="Yes" if d["is_rebalance_day"] else "")

            if d["is_rebalance_day"]:
                for c in range(1, len(headers2) + 1):
                    ws2.cell(row=row, column=c).fill = rebal_fill

            prev_row_map[sym] = row

        for c in range(1, len(headers2) + 1):
            ws2.column_dimensions[get_column_letter(c)].width = 16

        # ================================================================
        # Sheet 3 — Rebalance History
        # ================================================================
        ws3 = wb.create_sheet("Rebalances")
        headers3 = [
            "Rebalance Date", "Symbol", "Name", "Sector",
            "Weight %", "Direction Accuracy %",
        ]
        ws3.append(headers3)
        _style_header(ws3, len(headers3))
        r = 2
        for comp in compositions:
            for h in comp.get("holdings", []):
                ws3.cell(row=r, column=1, value=comp["date"])
                ws3.cell(row=r, column=2, value=h["symbol"])
                ws3.cell(row=r, column=3, value=h["name"])
                ws3.cell(row=r, column=4, value=h["sector"])
                ws3.cell(row=r, column=5, value=h["weight"])
                ws3.cell(row=r, column=5).number_format = '0.00'
                ws3.cell(row=r, column=6, value=h["accuracy"])
                ws3.cell(row=r, column=6).number_format = '0.00'
                r += 1
        for c in range(1, len(headers3) + 1):
            ws3.column_dimensions[get_column_letter(c)].width = 20

        # ================================================================
        # Sheet 4 — Summary Statistics
        # ================================================================
        ws4 = wb.create_sheet("Summary")
        ws4.append(["Metric", "Robo-Index", "SMI"])
        _style_header(ws4, 3)
        for label, key in [("Total Return %", "total_return"), ("Annualized Vol %", "annualized_vol"),
                           ("Sharpe Ratio", "sharpe_ratio"), ("Max Drawdown %", "max_drawdown"),
                           ("Final Value", "final_value")]:
            ws4.append([label, p_stats.get(key), s_stats.get(key)])
        ws4.append([])
        ws4.append(["Backtest Period (weeks)", weeks])
        ws4.append(["Rebalancing Frequency", {"D": "Daily", "W": "Weekly", "M": "Monthly", "3M": "Quarterly"}[rebal]])
        for c in range(1, 4):
            ws4.column_dimensions[get_column_letter(c)].width = 22

        # Auto-filter on main sheets
        ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{last_data}"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(detail)+1}"
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{r-1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"robo_index_report_{rebal}_{weeks}w_{datetime.date.today()}.xlsx"
        return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        import traceback
        print(f"Error in /api/robo-index/report: {e}\n{traceback.format_exc()}")
        return _json_response({"error": str(e)}), 500


@app.route("/api/predictions")
def api_predictions():
    symbol = request.args.get("symbol", default=None, type=str)
    days = request.args.get("days", default=None, type=int)
    start = request.args.get("start", default=None, type=str)
    end = request.args.get("end", default=None, type=str)
    predictor = _get_predictor(symbol)
    df, correct_direction_perc, mae, rmse, mape, close_correct, mae_pct, rmse_pct = predictor.fetch_prediction_history(
        limit=days, start_date=start, end_date=end
    )
    df["Date"] = df["Date"].astype(str)
    # Replace NaN/None so JSON serialization doesn't produce invalid tokens
    rows = json.loads(df.to_json(orient="records"))
    forecast = predictor.fetch_next_day_forecast()
    return _json_response({
        "correct_direction_pct": round(correct_direction_perc, 2),
        "mae": round(mae, 2),
        "mae_pct": round(mae_pct, 2),
        "rmse": round(rmse, 2),
        "rmse_pct": round(rmse_pct, 2),
        "mape": round(mape, 2),
        "close_correct": round(close_correct, 2),
        "rows": rows,
        "next_day": forecast,
    })


@app.route("/api/download")
def download_csv():
    predictor = _get_predictor()
    df, *_ = predictor.fetch_prediction_history()
    buf = io.BytesIO(df.to_csv(index=False).encode("utf-8"))
    filename = f"gold_prediction_report_{datetime.date.today()}.csv"
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name=filename)


@app.route("/api/store_predictions", methods=["POST"])
def store_predictions():
    body = request.get_json(silent=True) or {}
    symbol = body.get("symbol", SYMBOL)
    try:
        predictor = _get_predictor(symbol)
        forecast_df, past_df, df = predictor.create_predictions()
        predictor.store_predictions(past_df)
        predictor.store_predictions(forecast_df)
        predictor.update_with_real_close(df)
        # Calculate and store stats after storing predictions
        predictor.store_latest_stats()
        all_preds = pd.concat([past_df, forecast_df])
        min_date = str(all_preds["Date"].min())
        max_date = str(all_preds["Date"].max())
        total = len(all_preds)
        return _json_response({
            "message": f"Predictions for {symbol} stored successfully.",
            "min_date": min_date,
            "max_date": max_date,
            "total_rows": total,
        })
    except Exception as e:
        return _json_response({"error": f"Failed to create predictions for {symbol}: {str(e)}"}), 500


@app.route("/api/flush_predictions", methods=["POST"])
def flush_predictions():
    from sqlalchemy import create_engine, text as sa_text
    body = request.get_json(silent=True) or {}
    symbol = body.get("symbol", "").strip().upper()
    if not symbol:
        return _json_response({"error": "No symbol provided."}), 400
    engine = create_engine(DATABASE_URL)
    with engine.begin() as conn:
        result = conn.execute(
            sa_text("DELETE FROM predictions WHERE symbol = :symbol"),
            {"symbol": symbol},
        )
        deleted = result.rowcount
    return _json_response({"message": f"Deleted {deleted} rows for {symbol}.", "deleted": deleted})


@app.route("/robots.txt")
def robots():
    return send_file("robots.txt", mimetype="text/plain")


@app.route("/sitemap.xml")
def sitemap():
    return send_file("sitemap.xml", mimetype="application/xml")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)